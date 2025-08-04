import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 여러 파일을 인자로 받아서 각각 처리
if len(sys.argv) < 2:
    print('ERROR: 엑셀 파일 경로를 인자로 전달하세요.')
    sys.exit(1)

for excel_path in sys.argv[1:]:
    try:
        wb = load_workbook(excel_path)
        sheet_names = wb.sheetnames
        # Index 시트가 없으면 맨 앞에 생성
        if "Index" not in sheet_names:
            idx_sheet = wb.create_sheet("Index", 0)
        else:
            idx_sheet = wb["Index"]
            idx_sheet.delete_rows(1, idx_sheet.max_row)  # 기존 내용 삭제
        row = 1
        for name in sheet_names:
            if name == "Index":
                continue
            cell = idx_sheet.cell(row=row, column=1, value=name)
            # 하이퍼링크 추가
            cell.hyperlink = f"#{name}!A1"
            cell.style = "Hyperlink"
            row += 1
        wb.save(excel_path)
        print(f"{excel_path}: Index 시트 생성 완료")
    except Exception as e:
        print(f'ERROR: {excel_path}: {e}')
